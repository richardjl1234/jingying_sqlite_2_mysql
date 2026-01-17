"""
Load quota data from SQLite database, calculate obsolete_date for each record,
and export the results to MySQL database.

This program:
1. Reads quota data from SQLite database (quota table)
2. Calculates obsolete_date for each record based on effective dates
3. Maps data to MySQL quotas table using code dictionaries
4. Loads the processed data to MySQL database
"""

import os
import re
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from sql_util import sqlite_sql, mysql_sql


def calculate_obsolete_date(group):
    """
    Calculate obsolete_date for records within a group.
    
    For each group (类别1, 类别2, 型号, 加工工序):
    - First record: obsolete_date = next_record.effected_from - 1 day
    - Last record: obsolete_date = 99991231
    - Single record group: obsolete_date = 99991231
    
    Args:
        group: DataFrame subset for a specific group
        
    Returns:
        DataFrame with obsolete_date column added
    """
    # Sort by effected_from within the group
    group = group.sort_values('effected_from').reset_index(drop=True)
    
    # Initialize obsolete_date column
    group['obsolete_date'] = None
    
    if len(group) == 1:
        # Single record in group: obsolete_date = 99991231
        group.loc[0, 'obsolete_date'] = '99991231'
    else:
        # Multiple records in group
        for i in range(len(group) - 1):
            # Get next record's effected_from
            next_effected_from = group.loc[i + 1, 'effected_from']
            
            # Convert to datetime, subtract 1 day
            if isinstance(next_effected_from, str):
                next_date = datetime.strptime(str(next_effected_from), '%Y%m%d')
            else:
                next_date = pd.to_datetime(next_effected_from)
            
            # Subtract 1 day and format back to YYYYMMDD
            obsolete_date = next_date - timedelta(days=1)
            group.loc[i, 'obsolete_date'] = obsolete_date.strftime('%Y%m%d')
        
        # Last record: obsolete_date = 99991231
        group.loc[len(group) - 1, 'obsolete_date'] = '99991231'
    
    return group


def get_quota_with_obsolete_date():
    """
    Read quota data from SQLite database and calculate obsolete_date for each record.
    
    This function:
    1. Reads all columns from the quota table including '代码'
    2. Groups data by (类别1, 类别2, 型号, 加工工序)
    3. Calculates obsolete_date for each record within each group
    4. Returns the processed DataFrame sorted by 代码 to maintain original sequence
    
    Returns:
        pandas.DataFrame: Processed quota data with obsolete_date column
                          Sorted by 代码 to maintain original SQLite sequence
    """
    # Build SQL query to select all columns from quota table including 代码
    query = """
    SELECT 代码, 类别1, 类别2, 型号, 加工工序, 定额, effected_from 
    FROM quota 
    ORDER BY 代码
    """
    
    print("Reading quota data from SQLite database...")
    df = sqlite_sql(query)
    
    print(f"Retrieved {len(df)} records from quota table")
    print(f"Columns: {df.columns.tolist()}")
    
    # Check if we have any data
    if df.empty:
        print("No data found in quota table")
        return df
    
    # Print sample data
    print("\nSample data (first 5 records):")
    print(df.head())
    
    # Group by (类别1, 类别2, 型号, 加工工序) and calculate obsolete_date
    print("\nCalculating obsolete_date for each record...")
    
    group_columns = ['类别1', '类别2', '型号', '加工工序']
    
    # Apply calculate_obsolete_date to each group
    # Note: include_groups=True (default) keeps grouping columns in the result
    df_with_obsolete = df.groupby(group_columns, group_keys=False).apply(calculate_obsolete_date)
    
    # Suppress the FutureWarning about grouping columns (expected behavior for this use case)
    import warnings
    warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')
    
    # Reset index if it was added by groupby
    if 'level_0' in df_with_obsolete.columns:
        df_with_obsolete = df_with_obsolete.drop(columns=['level_0'])
    if 'index' in df_with_obsolete.columns:
        df_with_obsolete = df_with_obsolete.drop(columns=['index'])
    
    # Reset the index (groupby creates a multi-index)
    df_with_obsolete = df_with_obsolete.reset_index(drop=True)
    
    # Sort by the 代码 column to maintain original SQLite sequence
    df_with_obsolete = df_with_obsolete.sort_values(by='代码').reset_index(drop=True)
    
    print(f"Processed {len(df_with_obsolete)} records with obsolete_date")
    
    # Print sample of processed data
    print("\nSample processed data (first 10 records):")
    print(df_with_obsolete.head(10))
    
    return df_with_obsolete


def get_cat1_dict():
    """
    Get category 1 code dictionary from MySQL database.
    
    Returns:
        dict: Mapping of category 1 names to codes {name: cat1_code}
        
    Raises:
        ValueError: If duplicate names found in the table
    """
    print("Fetching category 1 codes from MySQL...")
    query = "SELECT cat1_code, name FROM process_cat1"
    df = mysql_sql(query)
    
    if df.empty:
        raise ValueError("No data found in process_cat1 table")
    
    # Check for duplicate names
    duplicate_names = df[df.duplicated(subset=['name'], keep=False)]['name'].unique()
    if len(duplicate_names) > 0:
        raise ValueError(f"Duplicate names found in process_cat1 table: {duplicate_names.tolist()}")
    
    # Create dictionary {name: cat1_code}
    cat1_dict = dict(zip(df['name'], df['cat1_code']))
    
    print(f"Loaded {len(cat1_dict)} category 1 codes")
    return cat1_dict


def get_cat2_dict():
    """
    Get category 2 code dictionary from MySQL database.
    
    Returns:
        dict: Mapping of category 2 names to codes {name: cat2_code}
        
    Raises:
        ValueError: If duplicate names found in the table
    """
    print("Fetching category 2 codes from MySQL...")
    query = "SELECT cat2_code, name FROM process_cat2"
    df = mysql_sql(query)
    
    if df.empty:
        raise ValueError("No data found in process_cat2 table")
    
    # Check for duplicate names
    duplicate_names = df[df.duplicated(subset=['name'], keep=False)]['name'].unique()
    if len(duplicate_names) > 0:
        raise ValueError(f"Duplicate names found in process_cat2 table: {duplicate_names.tolist()}")
    
    # Create dictionary {name: cat2_code}
    cat2_dict = dict(zip(df['name'], df['cat2_code']))
    
    print(f"Loaded {len(cat2_dict)} category 2 codes")
    return cat2_dict


def get_model_dict():
    """
    Get motor model code dictionary from MySQL database.
    
    Returns:
        dict: Mapping of model names to codes {name: model_code}
        
    Raises:
        ValueError: If duplicate names found in the table
    """
    print("Fetching motor model codes from MySQL...")
    query = "SELECT model_code, name FROM motor_models"
    df = mysql_sql(query)
    
    if df.empty:
        raise ValueError("No data found in motor_models table")
    
    # Check for duplicate names
    duplicate_names = df[df.duplicated(subset=['name'], keep=False)]['name'].unique()
    if len(duplicate_names) > 0:
        raise ValueError(f"Duplicate names found in motor_models table: {duplicate_names.tolist()}")
    
    # Create dictionary {name: model_code}
    model_dict = dict(zip(df['name'], df['model_code']))
    
    print(f"Loaded {len(model_dict)} motor model codes")
    return model_dict


def get_process_dict():
    """
    Get process code dictionary from MySQL database.
    
    Returns:
        dict: Mapping of process names to codes {name: process_code}
        
    Raises:
        ValueError: If duplicate names found in the table
    """
    print("Fetching process codes from MySQL...")
    query = "SELECT process_code, name FROM processes"
    df = mysql_sql(query)
    
    if df.empty:
        raise ValueError("No data found in processes table")
    
    # Check for duplicate names
    duplicate_names = df[df.duplicated(subset=['name'], keep=False)]['name'].unique()
    if len(duplicate_names) > 0:
        raise ValueError(f"Duplicate names found in processes table: {duplicate_names.tolist()}")
    
    # Create dictionary {name: process_code}
    process_dict = dict(zip(df['name'], df['process_code']))
    
    print(f"Loaded {len(process_dict)} process codes")
    return process_dict


def map_dataframe_to_quotas(df, cat1_dict, cat2_dict, model_dict, process_dict):
    """
    Map quota dataframe to MySQL quotas table format.
    
    Note: The '代码' column is used for sorting to maintain original SQLite sequence
    but is NOT loaded to MySQL (discarded after sorting).
    
    Column mapping:
    - 类别1 -> cat1_code (using cat1_dict)
    - 类别2 -> cat2_code (using cat2_dict)
    - 型号 -> model_code (using model_dict)
    - 加工工序 -> process_code (using process_dict)
    - 定额 -> unit_price (direct)
    - effected_from -> effective_date (direct)
    - obsolete_date -> obsolete_date (direct)
    - 1 -> created_by (constant)
    - current_timestamp -> created_at (generated)
    
    Args:
        df: Source DataFrame with columns [代码, 类别1, 类别2, 型号, 加工工序, 定额, effected_from, obsolete_date]
        cat1_dict: Dictionary mapping category 1 names to codes
        cat2_dict: Dictionary mapping category 2 names to codes
        model_dict: Dictionary mapping model names to codes
        process_dict: Dictionary mapping process names to codes
        
    Returns:
        pandas.DataFrame: Mapped data ready for MySQL insertion (without 代码 column)
        
    Raises:
        ValueError: If any mapping key is not found in the dictionary
    """
    print("\nMapping dataframe to MySQL quotas schema...")
    
    # Track missing keys for error reporting
    missing_cat1 = set()
    missing_cat2 = set()
    missing_model = set()
    missing_process = set()
    
    # Create new dataframe with MySQL column names
    quotas_df = pd.DataFrame()
    
    # Map 类别1 to cat1_code
    quotas_df['cat1_code'] = df['类别1'].apply(
        lambda x: cat1_dict.get(x) if x not in missing_cat1 else None
    )
    for val in df['类别1']:
        if val not in cat1_dict:
            missing_cat1.add(val)
    
    # Map 类别2 to cat2_code
    quotas_df['cat2_code'] = df['类别2'].apply(
        lambda x: cat2_dict.get(x) if x not in missing_cat2 else None
    )
    for val in df['类别2']:
        if val not in cat2_dict:
            missing_cat2.add(val)
    
    # Map 型号 to model_code
    quotas_df['model_code'] = df['型号'].apply(
        lambda x: model_dict.get(x) if x not in missing_model else None
    )
    for val in df['型号']:
        if val not in model_dict:
            missing_model.add(val)
    
    # Map 加工工序 to process_code
    quotas_df['process_code'] = df['加工工序'].apply(
        lambda x: process_dict.get(x) if x not in missing_process else None
    )
    for val in df['加工工序']:
        if val not in process_dict:
            missing_process.add(val)
    
    # Map other columns directly
    quotas_df['unit_price'] = df['定额']
    quotas_df['effective_date'] = df['effected_from'].astype(str)
    quotas_df['obsolete_date'] = df['obsolete_date'].astype(str)
    quotas_df['created_by'] = 1
    
    # Generate created_at timestamp
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    quotas_df['created_at'] = current_time
    
    # Check for missing keys and raise error if any
    errors = []
    if missing_cat1:
        errors.append(f"Missing cat1 codes for values: {sorted(missing_cat1)}")
    if missing_cat2:
        errors.append(f"Missing cat2 codes for values: {sorted(missing_cat2)}")
    if missing_model:
        errors.append(f"Missing model codes for values: {sorted(missing_model)}")
    if missing_process:
        errors.append(f"Missing process codes for values: {sorted(missing_process)}")
    
    if errors:
        raise ValueError("Mapping errors:\n" + "\n".join(errors))
    
    print(f"Mapped {len(quotas_df)} records to quotas schema")
    print(f"Output columns: {quotas_df.columns.tolist()}")
    
    # Print sample of mapped data
    print("\nSample mapped data (first 5 records):")
    print(quotas_df.head())
    
    return quotas_df


def sanitize_sheet_name(name):
    """
    Remove characters that are not suitable for Excel sheet names.
    
    Excel sheet names cannot contain: \ / ? * [ ] : and cannot be empty or longer than 31 characters.
    
    Args:
        name: The original sheet name
        
    Returns:
        str: Sanitized sheet name
    """
    # Remove invalid characters for Excel sheet names
    invalid_chars = r'[\\/\?\*\[\]:]'
    sanitized = re.sub(invalid_chars, '', name)
    # Remove leading/trailing whitespace
    sanitized = sanitized.strip()
    # Truncate to 31 characters if necessary
    if len(sanitized) > 31:
        sanitized = sanitized[:31]
    # If empty after sanitization, use a default name
    if not sanitized:
        sanitized = 'Sheet'
    return sanitized


def get_model_sort_key(model_code):
    """
    Extract the sort key from a model code.
    
    The sort key is the integer value of the part before the first hyphen.
    For example:
        "100-2" -> 100
        "63-1" -> 63
        "999" -> 999
    
    Args:
        model_code: The model code string
        
    Returns:
        int: The sort key
    """
    try:
        # Extract the part before the first hyphen
        part = model_code.split('-')[0]
        return int(part)
    except (ValueError, AttributeError):
        # If conversion fails, return 0
        return 0


def get_column_seq_dict():
    """
    Get column sequence from SQLite column_seq table.
    
    Returns:
        dict: Mapping of (类别1, 类别2, 加工工序) to seq number
    """
    print("Fetching column sequence from SQLite database...")
    query = "SELECT 类别1, 类别2, 加工工序, seq FROM column_seq"
    df = sqlite_sql(query)
    
    if df.empty:
        print("Warning: No data found in column_seq table, using default column order")
        return None
    
    # Create dictionary with (类别1, 类别2, 加工工序) as key and seq as value
    seq_dict = {}
    for _, row in df.iterrows():
        key = (row['类别1'], row['类别2'], row['加工工序'])
        seq_dict[key] = row['seq']
    
    print(f"Loaded {len(seq_dict)} column sequences")
    return seq_dict


def check_column_seq_table_exists():
    """
    Check if column_seq table exists in MySQL database.
    
    Returns:
        bool: True if table exists, False otherwise
    """
    print("Checking if column_seq table exists in MySQL...")
    query = "SHOW TABLES LIKE 'column_seq'"
    df = mysql_sql(query)
    
    # Handle case where mysql_sql returns None (e.g., connection issues or table doesn't exist)
    if df is None or df.empty:
        print("  column_seq table does not exist in MySQL")
        return False
    
    print("  column_seq table already exists in MySQL")
    return True


def create_column_seq_table():
    """
    Create column_seq table in MySQL if it doesn't exist.
    
    Table schema:
    - id INT AUTO_INCREMENT PRIMARY KEY
    - cat1_code VARCHAR(50) NOT NULL
    - cat2_code VARCHAR(50) NOT NULL
    - process_code VARCHAR(50) NOT NULL
    - seq INT NOT NULL
    """
    print("Creating column_seq table in MySQL...")
    
    create_table_sql = """
    CREATE TABLE IF NOT EXISTS column_seq (
        id INT AUTO_INCREMENT PRIMARY KEY,
        cat1_code VARCHAR(50) NOT NULL,
        cat2_code VARCHAR(50) NOT NULL,
        process_code VARCHAR(50) NOT NULL,
        seq INT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        INDEX idx_cat1_cat2 (cat1_code, cat2_code)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """
    
    mysql_sql(create_table_sql)
    print("  column_seq table created successfully (or already exists)")


def get_column_seq_from_sqlite():
    """
    Read all data from SQLite column_seq table.
    
    Returns:
        pandas.DataFrame: Data with columns [类别1, 类别2, 加工工序, seq]
    """
    print("Reading column_seq data from SQLite...")
    query = "SELECT 类别1, 类别2, 加工工序, seq FROM column_seq"
    df = sqlite_sql(query)
    
    if df.empty:
        print("  No data found in SQLite column_seq table")
    else:
        print(f"  Retrieved {len(df)} records from SQLite column_seq table")
    
    return df


def map_column_seq_to_mysql(df, cat1_dict, cat2_dict, process_dict):
    """
    Map column_seq DataFrame to MySQL column_seq schema.
    
    Column mapping:
    - 类别1 -> cat1_code (using cat1_dict)
    - 类别2 -> cat2_code (using cat2_dict)
    - 加工工序 -> process_code (using process_dict)
    - seq -> seq (direct)
    
    Args:
        df: Source DataFrame with columns [类别1, 类别2, 加工工序, seq]
        cat1_dict: Dictionary mapping category 1 names to codes
        cat2_dict: Dictionary mapping category 2 names to codes
        process_dict: Dictionary mapping process names to codes
        
    Returns:
        pandas.DataFrame: Mapped data ready for MySQL insertion
        
    Raises:
        ValueError: If any mapping key is not found in the dictionary
    """
    print("Mapping column_seq data to MySQL schema...")
    
    # Track missing keys for error reporting
    missing_cat1 = set()
    missing_cat2 = set()
    missing_process = set()
    
    # Create new dataframe with MySQL column names
    column_seq_df = pd.DataFrame()
    
    # Map 类别1 to cat1_code
    for val in df['类别1']:
        if val not in cat1_dict:
            missing_cat1.add(val)
    
    # Map 类别2 to cat2_code
    for val in df['类别2']:
        if val not in cat2_dict:
            missing_cat2.add(val)
    
    # Map 加工工序 to process_code
    for val in df['加工工序']:
        if val not in process_dict:
            missing_process.add(val)
    
    # Check for missing keys and raise error if any
    errors = []
    if missing_cat1:
        errors.append(f"Missing cat1 codes for values: {sorted(missing_cat1)}")
    if missing_cat2:
        errors.append(f"Missing cat2 codes for values: {sorted(missing_cat2)}")
    if missing_process:
        errors.append(f"Missing process codes for values: {sorted(missing_process)}")
    
    if errors:
        raise ValueError("Mapping errors:\n" + "\n".join(errors))
    
    # Apply mappings
    column_seq_df['cat1_code'] = df['类别1'].map(cat1_dict)
    column_seq_df['cat2_code'] = df['类别2'].map(cat2_dict)
    column_seq_df['process_code'] = df['加工工序'].map(process_dict)
    column_seq_df['seq'] = df['seq']
    
    print(f"  Mapped {len(column_seq_df)} records to column_seq schema")
    
    # Print sample of mapped data
    print("\n  Sample mapped data (first 5 records):")
    print(column_seq_df.head())
    
    return column_seq_df


def load_column_seq_to_mysql(column_seq_df):
    """
    Load column_seq DataFrame to MySQL database.
    
    Args:
        column_seq_df: DataFrame with MySQL column_seq table schema
        
    Returns:
        int: Number of records successfully loaded
    """
    table_name = "column_seq"
    
    print(f"Loading data to MySQL table: {table_name}")
    
    # Get MySQL database URL from environment variable
    MYSQL_DB_URL = os.environ.get("MYSQL_DB_URL")
    
    if not MYSQL_DB_URL:
        raise ValueError("MYSQL_DB_URL environment variable not set.")
    
    try:
        from sqlalchemy import create_engine
        engine = create_engine(MYSQL_DB_URL)
        
        # Clear existing data and load new data
        # First, check if there's existing data
        check_query = f"SELECT COUNT(*) as count FROM {table_name}"
        existing_df = mysql_sql(check_query)
        
        if not existing_df.empty and existing_df.iloc[0]['count'] > 0:
            print(f"  Clearing existing {existing_df.iloc[0]['count']} records from {table_name}")
            # Use truncate to clear the table (preserves auto_increment)
            mysql_sql(f"TRUNCATE TABLE {table_name}")
        
        # Deduplicate data before insert to avoid duplicate key errors
        # Keep the last occurrence of each unique (cat1_code, cat2_code, process_code) combination
        column_seq_df_deduped = column_seq_df.drop_duplicates(
            subset=['cat1_code', 'cat2_code', 'process_code'], 
            keep='last'
        )
        
        if len(column_seq_df_deduped) < len(column_seq_df):
            print(f"  Removed {len(column_seq_df) - len(column_seq_df_deduped)} duplicate records")
        
        # Bulk insert
        column_seq_df_deduped.to_sql(
            name=table_name,
            con=engine,
            if_exists='append',
            index=False,
            chunksize=100
        )
        
        print(f"  Successfully loaded {len(column_seq_df_deduped)} records to MySQL table: {table_name}")
        return len(column_seq_df)
        
    except Exception as e:
        raise ValueError(f"MySQL insert failed: {e}")


def get_process_sort_key(process_code, cat1, cat2, process_name, seq_dict):
    """
    Get the sort key for a process based on column_seq table.
    
    Args:
        process_code: The process code
        cat1: Category 1 name
        cat2: Category 2 name
        process_name: Process name
        seq_dict: Dictionary from get_column_seq_dict()
        
    Returns:
        int: Sequence number (or large number if not found)
    """
    if seq_dict is None:
        # Fallback to alphabetical sort
        return process_code
    
    # Try to find exact match
    key = (cat1, cat2, process_name)
    if key in seq_dict:
        return seq_dict[key]
    
    # If not found, try with code as name
    key = (cat1, cat2, process_code)
    if key in seq_dict:
        return seq_dict[key]
    
    # Not found, put at the end
    return float('inf')


def export_quota_to_excel(df, process_dict, cat1_dict, cat2_dict, model_dict, output_file_name, seq_dict=None):
    """
    Export quota data to an Excel file with multiple sheets.
    
    Each sheet corresponds to a unique combination of (cat1_code, effective_date).
    The table structure:
    - Column name: process name and code ("{process_name} ({process_code})")
    - Row index: model name and code ("{model_name} ({model_code})")
    - Cell values: unit_price
    
    Args:
        df: DataFrame containing quota data with columns including cat1_code, cat2_code, 
            model_code, process_code, unit_price, effective_date
        process_dict: Dictionary mapping process codes to process names {code: name}
        cat1_dict: Dictionary mapping cat1 codes to cat1 names {code: name}
        cat2_dict: Dictionary mapping cat2 codes to cat2 names {code: name}
        model_dict: Dictionary mapping model codes to model names {code: name}
        output_file_name: The output Excel file name (without path)
    """
    # Ensure output file has .xlsx extension
    if not output_file_name.endswith('.xlsx'):
        output_file_name += '.xlsx'
    
    print(f"\nExporting quota data to Excel: {output_file_name}")
    
    # Get all unique combinations of (cat1_code, effective_date)
    cat1_effective_pairs = df[['cat1_code', 'effective_date']].drop_duplicates()
    
    # Sort by cat1_code, then by effective_date
    cat1_effective_pairs = cat1_effective_pairs.sort_values(['cat1_code', 'effective_date'])
    
    # Create Excel writer with openpyxl engine
    with pd.ExcelWriter(output_file_name, engine='openpyxl') as writer:
        for _, row in cat1_effective_pairs.iterrows():
            cat1_code = row['cat1_code']
            effective_date = row['effective_date']
            
            # Format effective_date for sheet name
            if isinstance(effective_date, str):
                try:
                    date_obj = datetime.strptime(effective_date, '%Y%m%d')
                    formatted_date = date_obj.strftime('%Y-%m-%d')
                except ValueError:
                    formatted_date = effective_date
            else:
                formatted_date = pd.to_datetime(effective_date).strftime('%Y-%m-%d')
            
            # Create sheet name: "{cat1_name} {cat1_code} {effective_date}"
            cat1_name = cat1_dict.get(cat1_code, str(cat1_code))
            sheet_name = sanitize_sheet_name(f"{cat1_name} {cat1_code} {formatted_date}")
            
            print(f"  Processing sheet: {sheet_name}")
            
            # Filter dataframe for this cat1_code and effective_date
            sheet_df = df[(df['cat1_code'] == cat1_code) & (df['effective_date'] == effective_date)]
            
            # Get unique cat2 codes for this cat1
            unique_cat2_codes = sheet_df['cat2_code'].unique()
            
            # Track the current starting row for writing
            current_row = 0
            
            for cat2_code in sorted(unique_cat2_codes):
                # Filter dataframe for this cat1, cat2, and effective_date
                cat2_df = sheet_df[sheet_df['cat2_code'] == cat2_code]
                
                # Output the cat2 header: "{cat2_name} ({cat2_code})"
                cat2_name = cat2_dict.get(cat2_code, str(cat2_code))
                cat2_header = f"{cat2_name} ({cat2_code})"
                
                # Get unique process codes (columns) sorted by seq from column_seq table
                unique_process_codes = sorted(cat2_df['process_code'].unique(), key=lambda code: get_process_sort_key(code, cat1_name, cat2_name, process_dict.get(code, str(code)), seq_dict))
                
                # Create column names: "{process_name}\n({process_code})"
                column_names = []
                for process_code in unique_process_codes:
                    process_name = process_dict.get(process_code, str(process_code))
                    column_names.append(f"{process_name}\n({process_code})")
                
                # Get unique model codes (rows) and sort by the numeric part before hyphen
                unique_model_codes = sorted(cat2_df['model_code'].unique(), key=get_model_sort_key)
                
                # Create row index: "{model_name} ({model_code})"
                row_index = []
                for model_code in unique_model_codes:
                    model_name = model_dict.get(model_code, str(model_code))
                    row_index.append(f"{model_name} ({model_code})")
                
                # Build the data matrix
                data_dict = {col: [] for col in column_names}
                
                for model_code in unique_model_codes:
                    model_row_data = []
                    model_df = cat2_df[cat2_df['model_code'] == model_code]
                    
                    for process_code in unique_process_codes:
                        process_model_df = model_df[model_df['process_code'] == process_code]
                        
                        if not process_model_df.empty:
                            unit_price = process_model_df['unit_price'].values[0]
                        else:
                            unit_price = None
                        
                        model_row_data.append(unit_price)
                    
                    # Assign data to columns
                    for i, value in enumerate(model_row_data):
                        data_dict[column_names[i]].append(value)
                
                # Create the result dataframe
                result_df = pd.DataFrame(data_dict, index=row_index)
                
                # Write the cat2 header at current_row
                # First write the dataframe (which creates the sheet), then write the header
                result_df.index.name = '型号'
                result_df.to_excel(writer, sheet_name=sheet_name, startrow=current_row + 1, startcol=0, index=True, header=True)
                
                # Now access the sheet to write the cat2 header (sheet is now created)
                sheet = writer.sheets[sheet_name]
                
                # Apply styles
                # 1. Cat2 header: bold blue font
                cat2_cell = sheet.cell(row=current_row + 1, column=1, value=cat2_header)
                cat2_cell.font = Font(bold=True, color='0000FF')
                
                # 2. Column headers: light yellow background
                column_header_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                for col_idx in range(2, len(column_names) + 2):
                    header_cell = sheet.cell(row=current_row + 2, column=col_idx)
                    header_cell.fill = column_header_fill
                
                # Set row height for column header row to be 2x (30)
                sheet.row_dimensions[current_row + 2].height = 30
                
                # 3. Row index: light pink background
                row_index_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
                for row_idx in range(current_row + 3, current_row + 3 + len(row_index)):
                    index_cell = sheet.cell(row=row_idx, column=1)
                    index_cell.fill = row_index_fill
                
                # Auto-adjust column widths
                for col_idx in range(1, len(column_names) + 2):
                    max_width = 0
                    column_letter = get_column_letter(col_idx)
                    
                    # Check header row
                    if col_idx == 1:
                        header_text = '型号'
                    else:
                        header_text = column_names[col_idx - 2]
                    max_width = max(max_width, len(str(header_text)))
                    
                    # Check data rows
                    for row_idx in range(current_row + 3, current_row + 3 + len(row_index)):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value is not None:
                            max_width = max(max_width, len(str(cell_value)))
                    
                    # Add some padding
                    max_width = min(max_width * 1.3 + 5 , 50)  # Cap at 50
                    
                    # Set column width
                    sheet.column_dimensions[column_letter].width = max_width
                
                # Update current row for next cat2
                # +1 for cat2 header row, +1 for column header row, +len(result_df) for data rows, +2 for empty rows between tables
                current_row += len(result_df) + 4
    
    print(f"Excel file created successfully: {output_file_name}")


def load_to_mysql(quotas_df):
    """
    Load quotas DataFrame to MySQL database.
    
    Args:
        quotas_df: DataFrame with MySQL quotas table schema
        
    Returns:
        int: Number of records successfully loaded
    """
    table_name = "quotas"
    
    print(f"\nLoading data to MySQL table: {table_name}")
    
    # Get MySQL database URL from environment variable
    MYSQL_DB_URL = os.environ.get("MYSQL_DB_URL")
    
    if not MYSQL_DB_URL:
        raise ValueError("MYSQL_DB_URL environment variable not set.")
    
    try:
        from sqlalchemy import create_engine
        engine = create_engine(MYSQL_DB_URL)
        
        # Bulk insert with chunksize
        quotas_df.to_sql(
            name=table_name,
            con=engine,
            if_exists='append',
            index=False,
            chunksize=100
        )
        
        print(f"Successfully loaded {len(quotas_df)} records to MySQL table: {table_name}")
        return len(quotas_df)
        
    except Exception as e:
        raise ValueError(f"MySQL insert failed: {e}")


def main():
    """
    Main function to load quota data, calculate obsolete_date, map to MySQL schema, and load to database.
    """
    print("=" * 60)
    print("Load Quota to MySQL - Quota Processing and Loading Program")
    print("=" * 60)
    
    # Step 1: Get quota data with obsolete_date from SQLite
    print("\n[Step 1] Reading and processing quota data from SQLite...")
    df = get_quota_with_obsolete_date()
    
    if df.empty:
        print("No data to process. Exiting.")
        return
    
    # Step 2: Get code dictionaries from MySQL
    print("\n[Step 2] Fetching code dictionaries from MySQL...")
    
    try:
        cat1_dict = get_cat1_dict()
        cat2_dict = get_cat2_dict()
        model_dict = get_model_dict()
        process_dict = get_process_dict()
    except ValueError as e:
        print(f"Error fetching dictionaries: {e}")
        raise
    
    # Step 3: Map dataframe to MySQL quotas schema
    print("\n[Step 3] Mapping dataframe to MySQL quotas schema...")
    
    try:
        quotas_df = map_dataframe_to_quotas(df, cat1_dict, cat2_dict, model_dict, process_dict)
    except ValueError as e:
        print(f"Error mapping data: {e}")
        raise
    
    # Step 4: Load to MySQL database
    print("\n[Step 4] Loading data to MySQL database...")
    
    try:
        loaded_count = load_to_mysql(quotas_df)
    except ValueError as e:
        if "Duplicate entry" in str(e):
            print(f"  Warning: Data already exists in MySQL, skipping load...")
            loaded_count = 0
        else:
            print(f"Error loading to MySQL: {e}")
            raise
    
    # Step 5: Load column_seq table to MySQL
    print("\n[Step 5] Loading column_seq table to MySQL...")
    
    try:
        # Check if table exists and create if not
        if not check_column_seq_table_exists():
            print("  Creating column_seq table in MySQL...")
            create_column_seq_table()
        
        # Read column_seq data from SQLite
        column_seq_df = get_column_seq_from_sqlite()
        
        # Check if DataFrame is valid and not empty
        if column_seq_df is not None and not column_seq_df.empty:
            # Map to MySQL schema using dictionaries
            column_seq_mysql_df = map_column_seq_to_mysql(
                column_seq_df, cat1_dict, cat2_dict, process_dict
            )
            
            # Load to MySQL
            column_seq_loaded_count = load_column_seq_to_mysql(column_seq_mysql_df)
            print(f"  Loaded {column_seq_loaded_count} records to column_seq table")
        else:
            print("  No data in SQLite column_seq table to load")
            column_seq_loaded_count = 0
            
    except ValueError as e:
        print(f"  Warning: Error loading column_seq table: {e}")
        print("  Continuing with other operations...")
        column_seq_loaded_count = 0
    except Exception as e:
        print(f"  Warning: Unexpected error loading column_seq table: {e}")
        print("  Continuing with other operations...")
        column_seq_loaded_count = 0
    
    # Step 6: Export to Excel
    print("\n[Step 6] Exporting data to Excel...")
    
    try:
        # Create reversed dictionaries (code -> name) for export function
        cat1_dict_reversed = {v: k for k, v in cat1_dict.items()}
        cat2_dict_reversed = {v: k for k, v in cat2_dict.items()}
        model_dict_reversed = {v: k for k, v in model_dict.items()}
        process_dict_reversed = {v: k for k, v in process_dict.items()}
        
        # Get column sequence from SQLite for ordering columns
        seq_dict = get_column_seq_dict()
        
        export_quota_to_excel(quotas_df, process_dict_reversed, cat1_dict_reversed, cat2_dict_reversed, model_dict_reversed, "定额.xlsx", seq_dict)
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        raise
    
    # Display summary statistics
    print(f"\nSummary:")
    print(f"- Total records processed: {len(df)}")
    print(f"- Records loaded to MySQL (quotas): {loaded_count}")
    print(f"- Records loaded to MySQL (column_seq): {column_seq_loaded_count}")
    print(f"- Unique cat1 codes: {quotas_df['cat1_code'].nunique()}")
    print(f"- Unique effective dates: {quotas_df['effective_date'].nunique()}")
    print(f"- Unique cat2 codes: {quotas_df['cat2_code'].nunique()}")
    print(f"- Unique model codes: {quotas_df['model_code'].nunique()}")
    print(f"- Unique process codes: {quotas_df['process_code'].nunique()}")
    print(f"- Excel sheets: {quotas_df[['cat1_code', 'effective_date']].drop_duplicates().shape[0]}")
    print(f"- Excel file generated: 定额.xlsx")
    
    print(f"\nProcessing complete!")
    
    return quotas_df


if __name__ == "__main__":
    main()
